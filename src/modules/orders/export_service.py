"""Export service for generating PDF and Excel reports"""
from io import BytesIO
from datetime import date
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExportService:
    """Service for exporting data to PDF and Excel"""
    
    @staticmethod
    def generate_summary_pdf(stats: dict, low_stock: list, pending_summary: dict) -> BytesIO:
        """Generate PDF summary report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        # Container for elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1E40AF'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add title
        title = Paragraph("📊 Resumen del Sistema", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_RIGHT
        )
        date_text = Paragraph(f"Fecha: {stats.get('date', date.today().isoformat())}", date_style)
        elements.append(date_text)
        elements.append(Spacer(1, 0.3*inch))
        
        # Sales section
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#059669'),
            spaceAfter=12
        )
        
        sales_title = Paragraph("💰 Ventas del Día", section_style)
        elements.append(sales_title)
        
        # Sales data table
        sales_data = [
            ['Concepto', 'Valor'],
            ['Total Pedidos', str(stats.get('total_orders', 0))],
            ['Ventas Totales', f"${stats.get('total_sales', 0):.2f}"],
            ['Monto Cobrado', f"${stats.get('total_collected', 0):.2f}"],
            ['Cantidad de Pagos', str(stats.get('payments_count', 0))],
            ['Pedidos Pagados', str(stats.get('paid_orders', 0))],
            ['Pedidos Pendientes', str(stats.get('pending_orders', 0))]
        ]
        
        sales_table = Table(sales_data, colWidths=[3*inch, 2*inch])
        sales_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(sales_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Pending orders section
        pending_title = Paragraph("⏳ Pedidos Pendientes de Pago", section_style)
        elements.append(pending_title)
        
        pending_data = [
            ['Concepto', 'Valor'],
            ['Cantidad de Pedidos', str(pending_summary.get('count', 0))],
            ['Monto Total Pendiente', f"${pending_summary.get('total_amount', 0):.2f}"]
        ]
        
        pending_table = Table(pending_data, colWidths=[3*inch, 2*inch])
        pending_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D97706')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10)
        ]))
        
        elements.append(pending_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Low stock section
        low_stock_title = Paragraph(f"⚠️ Alertas de Stock Bajo ({len(low_stock)} productos)", section_style)
        elements.append(low_stock_title)
        
        if low_stock:
            stock_data = [['Producto', 'Stock', 'Mínimo', 'Precio']]
            for product in low_stock:
                stock_data.append([
                    product['nombre'],
                    str(product['stock']),
                    str(product['stock_minimo']),
                    f"${product['precio_venta']:.2f}"
                ])
            
            stock_table = Table(stock_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1.5*inch])
            stock_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(stock_table)
        else:
            no_alerts = Paragraph("No hay productos con stock bajo", styles['Normal'])
            elements.append(no_alerts)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_summary_excel(stats: dict, low_stock: list, pending_summary: dict) -> BytesIO:
        """Generate Excel summary report"""
        buffer = BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Styles
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Resumen General
        ws1 = wb.create_sheet("Resumen General")
        
        # Title
        ws1['A1'] = '📊 Resumen del Sistema'
        ws1['A1'].font = Font(bold=True, size=18, color="1E40AF")
        ws1['A1'].alignment = Alignment(horizontal='center')
        ws1.merge_cells('A1:D1')
        
        # Date
        ws1['A2'] = f"Fecha: {stats.get('date', date.today().isoformat())}"
        ws1['A2'].alignment = Alignment(horizontal='right')
        ws1.merge_cells('A2:D2')
        
        # Sales section
        row = 4
        ws1[f'A{row}'] = '💰 Ventas del Día'
        ws1[f'A{row}'].font = section_font
        ws1[f'A{row}'].fill = section_fill
        ws1.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws1[f'A{row}'] = 'Concepto'
        ws1[f'B{row}'] = 'Valor'
        ws1[f'A{row}'].font = header_font
        ws1[f'B{row}'].font = header_font
        ws1[f'A{row}'].fill = header_fill
        ws1[f'B{row}'].fill = header_fill
        
        sales_items = [
            ('Total Pedidos', stats.get('total_orders', 0)),
            ('Ventas Totales', f"${stats.get('total_sales', 0):.2f}"),
            ('Monto Cobrado', f"${stats.get('total_collected', 0):.2f}"),
            ('Cantidad de Pagos', stats.get('payments_count', 0)),
            ('Pedidos Pagados', stats.get('paid_orders', 0)),
            ('Pedidos Pendientes', stats.get('pending_orders', 0))
        ]
        
        for concept, value in sales_items:
            row += 1
            ws1[f'A{row}'] = concept
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].border = border
            ws1[f'B{row}'].border = border
        
        # Pending orders section
        row += 2
        ws1[f'A{row}'] = '⏳ Pedidos Pendientes de Pago'
        ws1[f'A{row}'].font = section_font
        ws1[f'A{row}'].fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        ws1.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws1[f'A{row}'] = 'Concepto'
        ws1[f'B{row}'] = 'Valor'
        ws1[f'A{row}'].font = header_font
        ws1[f'B{row}'].font = header_font
        ws1[f'A{row}'].fill = header_fill
        ws1[f'B{row}'].fill = header_fill
        
        pending_items = [
            ('Cantidad de Pedidos', pending_summary.get('count', 0)),
            ('Monto Total Pendiente', f"${pending_summary.get('total_amount', 0):.2f}")
        ]
        
        for concept, value in pending_items:
            row += 1
            ws1[f'A{row}'] = concept
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].border = border
            ws1[f'B{row}'].border = border
        
        # Adjust column widths
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20
        
        # Sheet 2: Stock Bajo
        ws2 = wb.create_sheet("Alertas de Stock")
        
        ws2['A1'] = f'⚠️ Productos con Stock Bajo ({len(low_stock)} productos)'
        ws2['A1'].font = Font(bold=True, size=14, color="DC2626")
        ws2.merge_cells('A1:D1')
        
        # Headers
        ws2['A2'] = 'Producto'
        ws2['B2'] = 'Stock Actual'
        ws2['C2'] = 'Stock Mínimo'
        ws2['D2'] = 'Precio Venta'
        
        for col in ['A2', 'B2', 'C2', 'D2']:
            ws2[col].font = header_font
            ws2[col].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            ws2[col].border = border
        
        # Data
        if low_stock:
            for idx, product in enumerate(low_stock, start=3):
                ws2[f'A{idx}'] = product['nombre']
                ws2[f'B{idx}'] = product['stock']
                ws2[f'C{idx}'] = product['stock_minimo']
                ws2[f'D{idx}'] = f"${product['precio_venta']:.2f}"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws2[f'{col}{idx}'].border = border
                    if product['stock'] == 0:
                        ws2[f'{col}{idx}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        else:
            ws2['A3'] = 'No hay productos con stock bajo'
            ws2.merge_cells('A3:D3')
        
        # Adjust column widths
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 15
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer
